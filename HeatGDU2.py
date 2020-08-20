#HeatGDU2.py
#To read heat on and around GDU
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import math
import numpy as np
# import dill
import unicodedata
import re
import csv
import os
import time
from kfxtools import * #fit for Python 2.


iHeat=load_workbook(filename='SCE_Heat_Template_GDU.xlsx')
shHeatMax = iHeat['Max']
shHeatAvg = iHeat['Avg']


iGDU=load_workbook(filename='Coord_GDU_packages.xlsx',data_only=True)
shCoords = iGDU['Sheet1']
iStart = 4
i = iStart
GDU = {}
while shCoords.cell(i,5).value != None:
    key = shCoords.cell(i,3).value    
    X = []
    for k in range(0,8):
        x = float(shCoords.cell(i,3*k+5).value)
        y = float(shCoords.cell(i,3*k+6).value)
        z = float(shCoords.cell(i,3*k+7).value)
        X.append([x,y,z])
    GDU[key] = X
    i += 1
nGDU = i - iStart


fieldnum = 0

NumRowsJetInfo = 5
basefolder = "./Rev.B/"
Js = {}
Js['J01'] = 'P04_A_AP'
Js['J02'] = 'P04_A_FP'
Js['J03'] = 'P04_A_AS'
Js['J04'] = 'P04_A_FS'
Js['J05'] = 'P04_B_AS'
Js['J06'] = 'P04_B_FS'
Js['J07'] = 'S05_A_AS'
Js['J08'] = 'S05_A_AP'
Js['J09'] = 'S05_A_F'
Js['J10'] = 'S05_A_F_10'
Js['J11'] = 'S05_B_AP'
Js['J12'] = 'S05_B_F'
Js['J13'] = 'S04_A_A'
Js['J14'] = 'S04_A_FP'
Js['J15'] = 'S04_A_FS'
Js['J16'] = 'S04_B_AP'
Js['J17'] = 'S04_B_F'
Js['J18'] = 'S03_A_P'
Js['J19'] = 'S03_B_A'
Js['J20'] = 'S03_B_F'
Js['J21'] = 'P02_B_FS'
Js['J22'] = 'P02_B_FP'
Js['J23'] = 'P05_A_P'
Js['J24'] = 'P05_A_S'
Js['J25'] = 'P05_B_A'
Js['J26'] = 'P05_B_F'
Js['J27'] = 'P03_B_P'
Js['J28'] = 'P03_B_FS'
Js['J29'] = 'KOD_B'

sce_box = open('GDU_box.kfx','w')
for j in Js.keys():
# for j in ['J02','J26']:
    colid = int(j[-2:])+10
    fdr = Js[j][:3]    
    fn = basefolder + "/" + j+"_rad_exit.r3d"    
    
    
    if (os.path.exists(fn) == False):
        print(fn + " does not exist")    
    else:    
        #Rad radiation
        T = readr3d(fn)            
        fieldname = T.names[fieldnum]
        print(fieldname)
        print(Js[j],fn,fieldname)
        rowid = 2
        for k in GDU.keys():        
        # for k in ['038-VZ-001']:        
            shHeatMax.cell(rowid+NumRowsJetInfo,1).value = k            
            shHeatAvg.cell(rowid+NumRowsJetInfo,1).value = k            
            X = np.array(GDU[k])
            x = min(X[:,0])
            dx = max(X[:,0])-x
            y = min(X[:,1])
            dy = max(X[:,1])-y
            z = min(X[:,2])
            dz = max(X[:,2])-z
                            
            Xs = [x, x + 0.5*dx,x + dx ]
            Ys = [y, y + 0.5*dy,y + dy ]
            Zs = [z, z + 0.5*dz,z + dz ]
            shHeatMax.cell(rowid+NumRowsJetInfo,4).value = Xs[1]            
            shHeatMax.cell(rowid+NumRowsJetInfo,5).value = Ys[1]            
            shHeatMax.cell(rowid+NumRowsJetInfo,6).value = Zs[1]            
            # shHeatAvg.cell(rowid+NumRowsJetInfo,1).value = k            
            r = 0.
            rmax = 0.
            rsum = 0.
            ncount = 0
            ravg = 0.
            for xi in Xs:
                for yi in Ys:                        
                    for zi in Zs:
                        r = T.point_value(xi,yi,zi,fieldnum)                                
                        # print("{:8.1f}{:8.1f}{:8.1f}{:8.1e}{:8.1f}".format(xi,yi,zi,r/1000,rmax/1000))
                        if r > 1000:
                            rmax = max(r,rmax)
                            ncount += 1
                            rsum += r
                    # ravg = rsum/(len(Xs)*len(Ys)*len(Zs)-1) #neglecting the reading at the center, that shall be 0!
                    
            if (ncount > 0) and (rsum > 0.01):
                ravg = rsum/ncount
                # print("#{:20s} {:6s} {:4s} {:6.1f} {:6.1f} {:10d}".format(k,s[6],s[7],rmax/1000,ravg/1000,ncount))
                sce_box.write("BOX: {:8.1f} {:8.1f} {:8.1f} {:8.1f} {:8.1f} {:8.1f} #{:15s}{:6.1f} {:6.1f} {:2d}\n".format((Xs[0])*1000,(Ys[0])*1000,(Zs[0])*1000,(Xs[2]-Xs[0])*1000,(Ys[2]-Ys[0])*1000,(Zs[2]-Zs[0])*1000,k,rmax/1000,ravg/1000,ncount))
                # shHeatMax.cell(rowid,colid).value = "{:8.2f}".format(rmax/1000)
                # shHeatAvg.cell(rowid,colid).value = "{:8.2f}".format(ravg/1000)                    
                shHeatMax.cell(rowid+NumRowsJetInfo,colid).value = rmax/1000
                shHeatAvg.cell(rowid+NumRowsJetInfo,colid).value = ravg/1000
            rowid += 1
               

iHeat.save('GDU_Heat_'+time.strftime("%Y%m%d-%H%M%S")+'.xlsx')
sce_box.close()