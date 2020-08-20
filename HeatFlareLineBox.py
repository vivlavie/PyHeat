#HeatFlareLine.py
#To read heat on points along the flare line
from kfxtools import * #fit for Python 2.
from openpyxl import load_workbook
import unicodedata
import re
import os
import time        

iHeat=load_workbook(filename='FlareLineCoord.xlsx',data_only=True)
shCoords = iHeat['20200812']
iStart = 4
i = iStart
Cs = {}
dx = 1
dy = 1
dz = 1
while shCoords.cell(i,2).value != None:
    key = shCoords.cell(i,2).value    
    x = float(shCoords.cell(i,8).value)
    y = float(shCoords.cell(i,9).value)
    z = float(shCoords.cell(i,10).value)
    Cs[key] = [i,x,y,z]
    i += 1
nCs = i - iStart

fieldnum = 0

basefolder = "./Rev.B/"
Js = ['J01','J02','J03','J04','J05','J06', 'J07','J08', 'J09','J10','J11','J12','J13','J14','J15','J16','J17','J18','J19','J20','J21','J22','J23','J24','J25','J26','J27','J28','J29']
# Js = ['J01', 'J02']
for j in Js:
# for j in :
    colid = int(j[-2:])+10           
    fn = basefolder + j+"_rad_exit.r3d"            
    if (os.path.exists(fn) == False):
        print(fn + " does not exist")    
    else:    
        #Rad radiation
        T = readr3d(fn)            
        fieldname = T.names[fieldnum]
        print(fieldname)
        print(j,fn,fieldname)

        #For each poitn along the flareline
        for c in Cs:
            rowid,x1,y1,z1 = Cs[c]


            Xs = [x1 - 1.0*dx, x1, x1 + 1.0*dx]
            Ys = [y1 - 1.0*dy, y1, y1 + 1.0*dy]
            Zs = [z1 - 1.0*dz, z1, z1 + 1.0*dz]
            r = 0.
            rmax = 0.
            rsum = 0.
            ncount = 0
            ravg = 0.
            for xi in Xs:
                for yi in Ys:                    
                    for zi in Zs:
                        r = T.point_value(xi,yi,zi,fieldnum)                                
                        # print("{:8.1f}{:8.1f}{:8.1f}{:8.1e}{:8.1f}".format(xi,yi,zi,r/1000,rmax/1000))
                        if r > 1000:
                            rmax = max(r,rmax)
                            ncount += 1
                            rsum += r
                    # ravg = rsum/(len(Xs)*len(Ys)*len(Zs)-1) #neglecting the reading at the center, that shall be 0!                    
            if (ncount > 0) and (rsum > 0.01):
                ravg = rsum/ncount
                # print("#{:20s} {:6s} {:4s} {:6.1f} {:6.1f} {:10d}".format(k,s[6],s[7],rmax/1000,ravg/1000,ncount))
                # sce_box.write("BOX: {:8.1f} {:8.1f} {:8.1f} {:8.1f} {:8.1f} {:8.1f} #{:15s} {:50s} {:6.1f} {:6.1f} {:2d}\n".format((Xs[0])*1000,(Ys[0])*1000,(Zs[0])*1000,(Xs[2]-Xs[0])*1000,(Ys[2]-Ys[0])*1000,(Zs[2]-Zs[0])*1000,k,sce_name,rmax/1000,ravg/1000,ncount))
                # shHeatMax.cell(rowid,colid).value = "{:8.2f}".format(rmax/1000)
                # shHeatAvg.cell(rowid,colid).value = "{:8.2f}".format(ravg/1000)
                shCoords.cell(rowid,colid).value = rmax/1000
                shCoords.cell(rowid+20,colid).value = ravg/1000
            else:
                shCoords.cell(rowid,colid).value = 'N/A'
                shCoords.cell(rowid+20,colid).value = 'N/A'


            
            #Impairment assessment w.r.t. Radiation
            # r1 = T.point_value(x1,y1,z1,fieldnum)     
            print(j,c,rowid, colid, rmax, ravg)    
            # shCoords.cell(rowid,colid).value = r1/1000
iHeat.save('HeatFlareLine_'+time.strftime("%Y%m%d-%H%M%S")+'.xlsx')


