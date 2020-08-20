#HeatFlareLine.py
#To read heat on points along the flare line
from kfxtools import * #fit for Python 2.
from openpyxl import load_workbook
import unicodedata
import re
import os
import time        

iHeat=load_workbook(filename='FlareLineCoord.xlsx',data_only=True)
shCoords = iHeat['sheet4']
iStart = 4
i = iStart
Cs = {}
while shCoords.cell(i,2).value != None:
    key = shCoords.cell(i,2).value    
    x = float(shCoords.cell(i,8).value)
    y = float(shCoords.cell(i,9).value)
    z = float(shCoords.cell(i,10).value)
    Cs[key] = [i,x,y,z]
    i += 1
nCs = i - iStart

fieldnum = 0

basefolder = "./Rev.B/"
# Js = ['J01','J02','J03','J04','J05','J06', 'J07','J08', 'J09','J10','J11','J12','J13','J14','J15','J16','J17','J18','J19','J20','J21','J22','J23','J24','J25','J26','J27','J28','J29']
Js = ['J01']
for j in Js:
# for j in :
    colid = int(j[-2:])+10           
    fn = basefolder + j+"_rad_exit.r3d"            
    if (os.path.exists(fn) == False):
        print(fn + " does not exist")    
    else:    
        #Rad radiation
        T = readr3d(fn)            
        fieldname = T.names[fieldnum]
        print(fieldname)
        print(j,fn,fieldname)

        #For each connection
        for c in Cs:
            rowid,x1,y1,z1 = Cs[c]
            
            #Impairment assessment w.r.t. Radiation
            r1 = T.point_value(x1,y1,z1,fieldnum)     
            print(j,c,r1)    
            shCoords.cell(rowid,colid).value = r1/1000
        iHeat.save('HeatFlareLine_'+time.strftime("%Y%m%d-%H%M%S")+'.xlsx')


